#!/usr/bin/env python3
"""Fast bulk import GMIM Excel data into Neon PostgreSQL"""
import glob, datetime, psycopg2

files = glob.glob('/home/z/my-project/upload/*.xlsx')
fpath = [f for f in files if '09.06' in f][0]
print(f"Dosya: {fpath.split('/')[-1]}")

import openpyxl
wb = openpyxl.load_workbook(fpath, data_only=True)

conn = psycopg2.connect('postgresql://neondb_owner:npg_Ycj6pC5wAJlb@ep-divine-dream-atgi8kfb-pooler.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require')
conn.autocommit = False
cur = conn.cursor()

# Clear
for t in ['Notification','ExamSupervisor','Exam','WeeklySchedule','PermanentDuty','Task','PointCategory','ImportLog','ResearchAssistant']:
    cur.execute(f'DELETE FROM "{t}"')
conn.commit()
print("🧹 Temizlendi")

# === AR.GÖR BİLGİLERİ ===
ws_info = wb['Ar.Gör. Bilgileri']
info_map = {}
for row in ws_info.iter_rows(min_row=2, max_row=ws_info.max_row, values_only=True):
    if row[2]:
        info_map[str(row[2]).strip()] = {
            'faculty': str(row[0]) if row[0] else 'DZ',
            'department': str(row[1]) if row[1] else 'GMI',
            'phone': str(row[3]) if row[3] else '',
            'email': str(row[4]) if row[4] else '',
        }

ws_toplam = wb['TOPLAM']
points_map = {}; order_list = []
for row in ws_toplam.iter_rows(min_row=2, max_row=12, values_only=True):
    if row[1]:
        name = str(row[1]).strip()
        points_map[name] = int(row[2]) if row[2] else 0
        order_list.append(name)

passwords = {'Begüm DOGANAY':'begum2026','Y.Tarık MUTLU':'tarik2026','Fatih NACAR':'fatih2026',
    'Samet BİÇEN':'samet2026','Merve GÜL ÇIVGIN':'merve2026','Sinan COŞKUN':'sinan2026',
    'Rukiye GÜLMEZ':'rukiye2026','Muhittin ORHAN':'muhittin2026','Cenk KAYA':'cenk2026','Ö. Berkehan İnal':'berkehan2026'}
admins = ['Begüm DOGANAY', 'Y.Tarık MUTLU']
inactive = ['Ö. Berkehan İnal', 'Muhittin ORHAN']

name_to_id = {}
for idx, name in enumerate(order_list):
    info = info_map.get(name, {})
    email = info.get('email', '')
    if not email:
        for key, val in info_map.items():
            if name.lower().replace(' ','').replace('.','').replace('ö','o').replace('ü','u') in key.lower().replace(' ','').replace('.','').replace('ö','o').replace('ü','u'):
                email = val.get('email', ''); break
    cur.execute('INSERT INTO "ResearchAssistant" (id,name,email,phone,faculty,department,"totalPoints","order","isActive",role,password,"createdAt","updatedAt") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW()) RETURNING id',
        (f'ra{idx}', name, email, info.get('phone') or None, info.get('faculty','DZ'), info.get('department','GMI'),
         points_map.get(name,0), idx+1, name not in inactive, 'admin' if name in admins else 'user', passwords.get(name,'argor2026')))
    name_to_id[name] = cur.fetchone()[0]
    st = "🔴 PASİF" if name in inactive else ("👑" if name in admins else "")
    print(f"  {name}: {points_map.get(name,0)}p {st}")
conn.commit()

# === PUAN BAREMİ ===
ws_barem = wb['Puan Baremi']
cat_name_to_id = {}; ci = 0
for row in ws_barem.iter_rows(min_row=1, max_row=55, values_only=True):
    nm = row[0]; pr = row[1] if len(row) > 1 else None
    if nm and pr:
        nm = str(nm).strip()
        if 'Puanlama' in nm: continue
        pts = 0
        if isinstance(pr, (int,float)): pts = int(pr)
        elif isinstance(pr, str):
            ns = [int(s) for s in pr.split() if s.isdigit()]
            if ns: pts = ns[0]
        if pts > 0:
            ci += 1
            cur.execute('INSERT INTO "PointCategory" (id,name,points,"isActive","createdAt","updatedAt") VALUES (%s,%s,%s,true,NOW(),NOW()) RETURNING id', (f'cat{ci}',nm,pts))
            cat_name_to_id[nm.lower().strip()] = cur.fetchone()[0]
conn.commit()
print(f"🏆 {ci} kategori")

# === KİŞİSEL GÖREVLER (BULK) ===
person_map = {'Begüm DOGANAY':'Begüm DOGANAY','Fatih NACAR':'Fatih NACAR','Y.Tarık MUTLU':'Y.Tarık MUTLU',
    'Merve GÜL ÇIVGIN':'Merve GÜL ÇIVGIN','Samet BİÇEN':'Samet BİÇEN','Sinan COŞKUN':'Sinan COŞKUN',
    'Rukiye GÜLMEZ':'Rukiye GÜLMEZ','Cenk KAYA':'Cenk KAYA','Berkehan İNAL':'Ö. Berkehan İnal','Muhittin ORHAN':'Muhittin ORHAN'}

total_tasks = 0
for sheet_name, person_name in person_map.items():
    if sheet_name not in wb.sheetnames: continue
    aid = name_to_id.get(person_name)
    if not aid: continue
    ws = wb[sheet_name]; tc = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        desc = str(row[1]).strip() if len(row)>1 and row[1] else ''
        if not desc: continue
        date_val = row[3] if len(row)>3 else None
        if isinstance(date_val, datetime.datetime): pass
        elif isinstance(date_val, str):
            try:
                p = date_val.split('.'); date_val = datetime.datetime(int(p[2]),int(p[1]),int(p[0]))
            except: date_val = datetime.datetime(2025,7,1)
        else: date_val = datetime.datetime(2025,7,1)
        p = int(row[4]) if len(row)>4 and isinstance(row[4],(int,float)) and row[4] else 0
        h = str(row[2]) if len(row)>2 and row[2] else None
        num = int(row[0]) if isinstance(row[0],(int,float)) and row[0] else tc+1
        cat_id = None
        dl = desc.lower()
        for cn, cid in cat_name_to_id.items():
            if cn in dl: cat_id = cid; break
        total_tasks += 1; tc += 1
        cur.execute('INSERT INTO "Task" (id,number,description,"hoursWorked",date,points,status,source,"assistantId","categoryId","createdAt","updatedAt") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())',
            (f't{total_tasks}', num, desc, h, date_val, p, 'approved', 'import', aid, cat_id))
    print(f"  {sheet_name}: {tc} görev")
conn.commit()

# === DAİMİ GÖREVLER ===
ws_genel = wb['Genel Görevler']
cp_id = None; do = 0; di = 0
for row in ws_genel.iter_rows(min_row=2, max_row=ws_genel.max_row, values_only=True):
    nr = row[0] if len(row)>0 else None
    d = row[1] if len(row)>1 else None
    if nr and str(nr).strip():
        nm = str(nr).strip()
        for n, aid in name_to_id.items():
            if nm.lower().replace(' ','') in n.lower().replace(' ','').replace('.',''): cp_id = aid; do = 0; break
    if d and cp_id:
        ds = str(d).strip()
        if ds and not ds.startswith('Ad-') and not ds.startswith('Görevi'):
            do += 1; di += 1
            cur.execute('INSERT INTO "PermanentDuty" (id,name,"order","assistantId","createdAt","updatedAt") VALUES (%s,%s,%s,%s,NOW(),NOW())', (f'd{di}',ds,do,cp_id))
conn.commit()
print(f"📌 Daimi görevler eklendi")

# === HAFTALIK PROGRAM ===
ws_prog = wb['Haftalık Program']
day_map = {'PAZARTESİ':1,'SALI':2,'ÇARŞAMBA':3,'PERŞEMBE':4,'CUMA':5}
day_names = list(day_map.keys()); si = 0
for row in ws_prog.iter_rows(min_row=3, max_row=ws_prog.max_row, values_only=False):
    nm = row[0].value
    if not nm: continue
    aid = None
    for n, a in name_to_id.items():
        if str(nm).strip().lower().replace(' ','') in n.lower().replace(' ','').replace('.',''): aid = a; break
    if not aid: continue
    for ci, dn in enumerate(day_names):
        cv = row[ci+1].value if ci+1 < len(row) else None
        if cv and str(cv).strip():
            ds = str(cv).strip(); si += 1
            cur.execute('INSERT INTO "WeeklySchedule" (id,"dayOfWeek","timeSlot",description,"assistantId","createdAt","updatedAt") VALUES (%s,%s,%s,%s,%s,NOW(),NOW())', (f's{si}',day_map[dn],ds,ds,aid))
conn.commit()
print(f"📅 Program eklendi")

# Summary
cur.execute('SELECT COUNT(*) FROM "ResearchAssistant"'); print(f"\n✅ Ar.Gör: {cur.fetchone()[0]}")
cur.execute('SELECT COUNT(*) FROM "Task"'); print(f"✅ Görevler: {cur.fetchone()[0]}")
cur.execute('SELECT COUNT(*) FROM "PointCategory"'); print(f"✅ Kategoriler: {cur.fetchone()[0]}")
cur.execute('SELECT COUNT(*) FROM "PermanentDuty"'); print(f"✅ Daimi: {cur.fetchone()[0]}")
cur.execute('SELECT COUNT(*) FROM "WeeklySchedule"'); print(f"✅ Program: {cur.fetchone()[0]}")
cur.execute('SELECT name,"isActive" FROM "ResearchAssistant" WHERE "isActive"=false'); 
for r in cur.fetchall(): print(f"🔴 PASİF: {r[0]}")
cur.close(); conn.close()
print("\n🎉 AKTARIM TAMAMLANDI!")
