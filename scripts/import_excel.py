#!/usr/bin/env python3
"""Import GMIM Excel data into Neon PostgreSQL database"""

import glob
import datetime
import asyncio
import os

# Set env vars before importing prisma

if not os.environ.get('DATABASE_URL'):
    raise RuntimeError('DATABASE_URL environment variable is required')

import openpyxl

# Use prisma via subprocess for DB operations
import subprocess
import json

def prisma_query(sql):
    """Execute SQL via prisma db execute"""
    result = subprocess.run(
        ['npx', 'prisma', 'db', 'execute', '--stdin'],
        input=sql,
        capture_output=True, text=True,
        env={**os.environ}
    )
    return result

def get_neon_connection():
    """Get psycopg2-style connection to Neon"""
    import psycopg2
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    return conn

def main():
    # Find Excel file
    files = glob.glob('/home/z/my-project/upload/*.xlsx')
    fpath = [f for f in files if '09.06' in f][0]
    print(f"Dosya: {fpath.split('/')[-1]}")
    
    wb = openpyxl.load_workbook(fpath, data_only=True)
    
    # Connect to Neon
    conn = get_neon_connection()
    cur = conn.cursor()
    
    # ==========================================
    # 1. CLEAR EXISTING DATA
    # ==========================================
    print("\n🧹 Mevcut veriler temizleniyor...")
    tables = ['Notification', 'ExamSupervisor', 'Exam', 'WeeklySchedule', 
              'PermanentDuty', 'Task', 'PointCategory', 'ImportLog', 'ResearchAssistant']
    for t in tables:
        cur.execute(f'DELETE FROM "{t}"')
    conn.commit()
    print("  Temizlendi!")
    
    # ==========================================
    # 2. AR.GÖR BİLGİLERİ
    # ==========================================
    print("\n👤 Araş görler ekleniyor...")
    ws_info = wb['Ar.Gör. Bilgileri']
    
    info_map = {}
    for row in ws_info.iter_rows(min_row=2, max_row=ws_info.max_row, values_only=True):
        if row[2]:
            name = str(row[2]).strip()
            info_map[name] = {
                'faculty': str(row[0]) if row[0] else 'DZ',
                'department': str(row[1]) if row[1] else 'GMI',
                'phone': str(row[3]) if row[3] else '',
                'email': str(row[4]) if row[4] else '',
            }
    
    # TOPLAM - points and order
    ws_toplam = wb['TOPLAM']
    points_map = {}
    order_list = []
    for row in ws_toplam.iter_rows(min_row=2, max_row=12, values_only=True):
        if row[1]:
            name = str(row[1]).strip()
            points_map[name] = int(row[2]) if row[2] else 0
            order_list.append(name)
    
    passwords = {
        'Begüm DOGANAY': 'begum2026',
        'Y.Tarık MUTLU': 'tarik2026',
        'Fatih NACAR': 'fatih2026',
        'Samet BİÇEN': 'samet2026',
        'Merve GÜL ÇIVGIN': 'merve2026',
        'Sinan COŞKUN': 'sinan2026',
        'Rukiye GÜLMEZ': 'rukiye2026',
        'Muhittin ORHAN': 'muhittin2026',
        'Cenk KAYA': 'cenk2026',
        'Ö. Berkehan İnal': 'berkehan2026',
    }
    
    admins = ['Begüm DOGANAY', 'Y.Tarık MUTLU']
    inactive = ['Ö. Berkehan İnal', 'Muhittin ORHAN']
    
    name_to_id = {}
    for idx, name in enumerate(order_list):
        info = info_map.get(name, {})
        email = info.get('email', '')
        if not email:
            for key, val in info_map.items():
                if name.lower().replace(' ', '').replace('.','').replace('ö','o').replace('ü','u') in key.lower().replace(' ', '').replace('.','').replace('ö','o').replace('ü','u'):
                    email = val.get('email', '')
                    break
        
        is_active = name not in inactive
        role = 'admin' if name in admins else 'user'
        password = passwords.get(name, 'argor2026')
        phone = info.get('phone', '')
        
        cur.execute("""
            INSERT INTO "ResearchAssistant" (id, name, email, phone, faculty, department, "totalPoints", "order", "isActive", role, password, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            RETURNING id
        """, (
            f'ra_{idx+1}', name, email, phone or None,
            info.get('faculty', 'DZ'), info.get('department', 'GMI'),
            points_map.get(name, 0), idx + 1,
            is_active, role, password
        ))
        ra_id = cur.fetchone()[0]
        name_to_id[name] = ra_id
        status = "🔴 PASİF" if not is_active else ("👑 Admin" if role == 'admin' else "")
        print(f"  {name}: {points_map.get(name, 0)} puan {status}")
    
    conn.commit()
    
    # ==========================================
    # 3. PUAN BAREMİ
    # ==========================================
    print("\n🏆 Puan baremi ekleniyor...")
    ws_barem = wb['Puan Baremi']
    cat_name_to_id = {}
    cat_idx = 0
    
    for row in ws_barem.iter_rows(min_row=1, max_row=55, values_only=True):
        name = row[0]
        points_raw = row[1] if len(row) > 1 else None
        
        if name and points_raw:
            name = str(name).strip()
            if 'Puanlama Sistemi' in name:
                continue
            
            points = 0
            if isinstance(points_raw, (int, float)):
                points = int(points_raw)
            elif isinstance(points_raw, str):
                nums = [int(s) for s in points_raw.split() if s.isdigit()]
                if nums:
                    points = nums[0]
            
            if points > 0 and name:
                cat_idx += 1
                cur.execute("""
                    INSERT INTO "PointCategory" (id, name, points, "isActive", "createdAt", "updatedAt")
                    VALUES (%s, %s, %s, true, NOW(), NOW())
                    RETURNING id
                """, (f'cat_{cat_idx}', name, points))
                cat_id = cur.fetchone()[0]
                cat_name_to_id[name.lower().strip()] = cat_id
                print(f"  {name}: {points} puan")
    
    conn.commit()
    
    # ==========================================
    # 4. KİŞİSEL GÖREVLER
    # ==========================================
    print("\n📋 Kişisel görevler aktarılıyor...")
    
    person_sheets = {
        'Begüm DOGANAY': 'Begüm DOGANAY',
        'Fatih NACAR': 'Fatih NACAR',
        'Y.Tarık MUTLU': 'Y.Tarık MUTLU',
        'Merve GÜL ÇIVGIN': 'Merve GÜL ÇIVGIN',
        'Samet BİÇEN': 'Samet BİÇEN',
        'Sinan COŞKUN': 'Sinan COŞKUN',
        'Rukiye GÜLMEZ': 'Rukiye GÜLMEZ',
        'Cenk KAYA': 'Cenk KAYA',
        'Berkehan İNAL': 'Ö. Berkehan İnal',
        'Muhittin ORHAN': 'Muhittin ORHAN',
    }
    
    total_tasks = 0
    
    for sheet_name, person_name in person_sheets.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ {sheet_name} sayfası bulunamadı")
            continue
        
        assistant_id = name_to_id.get(person_name)
        if not assistant_id:
            print(f"  ⚠️ {person_name} için ID bulunamadı")
            continue
        
        ws = wb[sheet_name]
        task_count = 0
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            num = row[0]
            desc = row[1] if len(row) > 1 else None
            hours = row[2] if len(row) > 2 else None
            date = row[3] if len(row) > 3 else None
            points = row[4] if len(row) > 4 else None
            
            if not desc and not date:
                continue
            
            desc = str(desc).strip() if desc else ''
            if not desc:
                continue
            
            # Parse date
            date_val = None
            if isinstance(date, datetime.datetime):
                date_val = date
            elif isinstance(date, str):
                try:
                    parts = date.split('.')
                    if len(parts) == 3:
                        date_val = datetime.datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                except:
                    pass
            
            if not date_val:
                date_val = datetime.datetime(2025, 7, 1)
            
            p = int(points) if isinstance(points, (int, float)) and points else 0
            h = str(hours) if hours else None
            
            # Find category
            cat_id = None
            desc_lower = desc.lower()
            for cat_name, cid in cat_name_to_id.items():
                if cat_name in desc_lower:
                    cat_id = cid
                    break
            
            try:
                cur.execute("""
                    INSERT INTO "Task" (id, number, description, "hoursWorked", date, points, status, source, "assistantId", "categoryId", "createdAt", "updatedAt")
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """, (
                    f'task_{total_tasks + 1}', 
                    int(num) if isinstance(num, (int, float)) and num else task_count + 1,
                    desc, h, date_val, p, 'approved', 'import',
                    assistant_id, cat_id
                ))
                task_count += 1
                total_tasks += 1
            except Exception as e:
                print(f"    ⚠️ Hata: {e}")
        
        print(f"  {sheet_name}: {task_count} görev")
    
    conn.commit()
    
    # ==========================================
    # 5. DAİMİ GÖREVLER
    # ==========================================
    print("\n📌 Daimi görevler aktarılıyor...")
    ws_genel = wb['Genel Görevler']
    
    current_person_id = None
    duty_order = 0
    duty_idx = 0
    
    for row in ws_genel.iter_rows(min_row=2, max_row=ws_genel.max_row, values_only=True):
        name_raw = row[0] if len(row) > 0 else None
        duty = row[1] if len(row) > 1 else None
        
        if name_raw and str(name_raw).strip():
            name = str(name_raw).strip()
            for n, aid in name_to_id.items():
                if name.lower().replace(' ', '') in n.lower().replace(' ', '').replace('.', ''):
                    current_person_id = aid
                    duty_order = 0
                    break
        
        if duty and current_person_id:
            duty_str = str(duty).strip()
            if duty_str and not duty_str.startswith('Ad-Soyadı') and not duty_str.startswith('Görevi'):
                duty_order += 1
                duty_idx += 1
                try:
                    cur.execute("""
                        INSERT INTO "PermanentDuty" (id, name, "order", "assistantId", "createdAt", "updatedAt")
                        VALUES (%s, %s, %s, %s, NOW(), NOW())
                    """, (f'duty_{duty_idx}', duty_str, duty_order, current_person_id))
                except Exception as e:
                    print(f"  ⚠️ Daimi görev hatası: {e}")
    
    conn.commit()
    print(f"  Daimi görevler eklendi")
    
    # ==========================================
    # 6. HAFTALIK PROGRAM
    # ==========================================
    print("\n📅 Haftalık program aktarılıyor...")
    ws_prog = wb['Haftalık Program']
    
    day_map = {'PAZARTESİ': 1, 'SALI': 2, 'ÇARŞAMBA': 3, 'PERŞEMBE': 4, 'CUMA': 5}
    day_names = list(day_map.keys())
    sched_idx = 0
    
    for row in ws_prog.iter_rows(min_row=3, max_row=ws_prog.max_row, values_only=False):
        name_cell = row[0].value
        if not name_cell:
            continue
        name = str(name_cell).strip()
        
        assistant_id = None
        for n, aid in name_to_id.items():
            if name.lower().replace(' ', '') in n.lower().replace(' ', '').replace('.', ''):
                assistant_id = aid
                break
        
        if not assistant_id:
            continue
        
        for col_idx, day_name in enumerate(day_names):
            cell_val = row[col_idx + 1].value if col_idx + 1 < len(row) else None
            if cell_val and str(cell_val).strip():
                desc = str(cell_val).strip()
                sched_idx += 1
                try:
                    cur.execute("""
                        INSERT INTO "WeeklySchedule" (id, "dayOfWeek", "timeSlot", description, "assistantId", "createdAt", "updatedAt")
                        VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                    """, (f'sched_{sched_idx}', day_map[day_name], desc, desc, assistant_id))
                except Exception as e:
                    print(f"  ⚠️ Program hatası: {e}")
    
    conn.commit()
    print(f"  Haftalık program eklendi")
    
    # ==========================================
    # SUMMARY
    # ==========================================
    cur.execute('SELECT COUNT(*) FROM "ResearchAssistant"')
    ra_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Task"')
    task_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "PointCategory"')
    cat_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "PermanentDuty"')
    duty_count = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "WeeklySchedule"')
    sched_count = cur.fetchone()[0]
    
    print(f"\n✅ AKTARIM TAMAMLANDI!")
    print(f"  Araş Gör: {ra_count}")
    print(f"  Görevler: {task_count}")
    print(f"  Kategoriler: {cat_count}")
    print(f"  Daimi Görevler: {duty_count}")
    print(f"  Haftalık Program: {sched_count}")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
